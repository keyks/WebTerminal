# backend/inspection.py
"""
集群状态巡检引擎
"""
import time
import uuid
import threading
from datetime import datetime
from typing import List, Optional, Callable
from dataclasses import dataclass, field, asdict

import openpyxl
from openpyxl.styles import Font, PatternFill

from logger import log

# ✅ 不在顶部导入 app，避免循环依赖


@dataclass
class NodeHealthStatus:
    node_id:          str
    host:             str
    name:             str
    status:           str
    cpu_percent:      float
    memory_percent:   float
    disk_percent:     float
    load_avg:         List[float]
    uptime:           str
    system:           str
    kernel:           str
    processes:        int
    network_rx:       int
    network_tx:       int
    is_abnormal:      bool = False
    abnormal_reasons: List[str] = field(default_factory=list)
    checked_at:       str = field(default_factory=lambda: datetime.now().isoformat())


@dataclass
class InspectionReport:
    report_id:      str
    title:          str
    generated_at:   str
    total_nodes:    int
    online_nodes:   int
    offline_nodes:  int
    abnormal_nodes: int
    nodes:          List[NodeHealthStatus] = field(default_factory=list)
    avg_cpu:        float = 0
    avg_memory:     float = 0
    avg_disk:       float = 0
    max_cpu:        float = 0
    max_memory:     float = 0
    max_disk:       float = 0


class InspectionEngine:
    def __init__(self):
        self._current_report: Optional[InspectionReport] = None

    def run_inspection(self, node_ids: List[str],
                       on_progress: Optional[Callable] = None) -> InspectionReport:
        # ✅ 延迟导入，避免循环依赖
        from core import db, ssh_manager

        report_id = f"inspect_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        report = InspectionReport(
            report_id=report_id,
            title=f"集群巡检报告 - {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            generated_at=datetime.now().isoformat(),
            total_nodes=len(node_ids),
            online_nodes=0,
            offline_nodes=0,
            abnormal_nodes=0
        )

        for idx, node_id in enumerate(node_ids):
            status = self._inspect_node(node_id, db, ssh_manager)
            if status:
                report.nodes.append(status)
                if status.status == 'online':
                    report.online_nodes += 1
                    if status.is_abnormal:
                        report.abnormal_nodes += 1
                else:
                    report.offline_nodes += 1

            if on_progress:
                try:
                    on_progress(idx + 1, len(node_ids))
                except Exception:
                    pass

        online_nodes = [n for n in report.nodes if n.status == 'online']
        if online_nodes:
            report.avg_cpu    = sum(n.cpu_percent    for n in online_nodes) / len(online_nodes)
            report.avg_memory = sum(n.memory_percent for n in online_nodes) / len(online_nodes)
            report.avg_disk   = sum(n.disk_percent   for n in online_nodes) / len(online_nodes)
            report.max_cpu    = max(n.cpu_percent    for n in online_nodes)
            report.max_memory = max(n.memory_percent for n in online_nodes)
            report.max_disk   = max(n.disk_percent   for n in online_nodes)

        return report

    def _inspect_node(self, node_id: str, db, ssh_manager) -> Optional[NodeHealthStatus]:
        conn = db.get_connection(node_id)
        if not conn:
            return NodeHealthStatus(
                node_id=node_id, host='unknown', name='unknown',
                status='error', cpu_percent=0, memory_percent=0, disk_percent=0,
                load_avg=[0, 0, 0], uptime='-', system='-', kernel='-',
                processes=0, network_rx=0, network_tx=0,
                is_abnormal=True, abnormal_reasons=['连接配置不存在']
            )

        session_id = f"inspect_{uuid.uuid4().hex[:8]}"
        success, msg = ssh_manager.create_session(
            session_id, conn, f"inspect-{conn.get('host')}"
        )
        if not success:
            return NodeHealthStatus(
                node_id=node_id,
                host=conn.get('host', 'unknown'),
                name=conn.get('name', 'unknown'),
                status='offline', cpu_percent=0, memory_percent=0, disk_percent=0,
                load_avg=[0, 0, 0], uptime='-', system='-', kernel='-',
                processes=0, network_rx=0, network_tx=0,
                is_abnormal=True, abnormal_reasons=[f'连接失败: {msg}']
            )

        session = ssh_manager.get_session(session_id)
        try:
            info = session.get_system_info()
            status = NodeHealthStatus(
                node_id=node_id,
                host=conn.get('host', 'unknown'),
                name=conn.get('name', 'unknown'),
                status='online',
                cpu_percent=info.get('cpu_percent', 0),
                memory_percent=info.get('memory', {}).get('percent', 0),
                disk_percent=info.get('disk', {}).get('percent', 0),
                load_avg=[
                    info.get('load', {}).get('load1',  0),
                    info.get('load', {}).get('load5',  0),
                    info.get('load', {}).get('load15', 0),
                ],
                uptime=info.get('uptime', '-'),
                system=info.get('system', '-'),
                kernel=info.get('kernel', '-'),
                processes=info.get('process_count', 0),
                network_rx=info.get('network', {}).get('rx_bytes', 0),
                network_tx=info.get('network', {}).get('tx_bytes', 0)
            )

            reasons = []
            if status.cpu_percent    > 85: reasons.append(f'CPU过高: {status.cpu_percent}%')
            if status.memory_percent > 85: reasons.append(f'内存过高: {status.memory_percent}%')
            if status.disk_percent   > 85: reasons.append(f'磁盘过高: {status.disk_percent}%')
            if status.load_avg[0]    > 10: reasons.append(f'负载过高: {status.load_avg[0]}')

            if reasons:
                status.is_abnormal      = True
                status.abnormal_reasons = reasons

            return status
        finally:
            ssh_manager.remove_session(session_id)

    def export_to_excel(self, report: InspectionReport, output_path: str):
        wb       = openpyxl.Workbook()
        ws       = wb.active
        ws.title = '汇总'

        for row, data in enumerate([
            ['巡检报告',  report.title],
            ['生成时间',  report.generated_at],
            ['总节点数',  report.total_nodes],
            ['在线节点',  report.online_nodes],
            ['离线节点',  report.offline_nodes],
            ['异常节点',  report.abnormal_nodes],
            ['平均CPU',  f"{report.avg_cpu:.1f}%"],
            ['平均内存',  f"{report.avg_memory:.1f}%"],
            ['平均磁盘',  f"{report.avg_disk:.1f}%"],
            ['最高CPU',  f"{report.max_cpu:.1f}%"],
            ['最高内存',  f"{report.max_memory:.1f}%"],
            ['最高磁盘',  f"{report.max_disk:.1f}%"],
        ], 1):
            for col, val in enumerate(data, 1):
                ws.cell(row=row, column=col, value=val)

        ws2     = wb.create_sheet('节点详情')
        headers = ['节点ID', '主机', '名称', '状态', 'CPU%', '内存%', '磁盘%',
                   '负载(1m)', '负载(5m)', '负载(15m)', '运行时间', '系统',
                   '进程数', '异常', '异常原因']
        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')
        yel_fill = PatternFill(start_color='FFD93D', end_color='FFD93D', fill_type='solid')

        for col, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)

        for row, node in enumerate(report.nodes, 2):
            vals = [
                node.node_id, node.host, node.name, node.status,
                node.cpu_percent, node.memory_percent, node.disk_percent,
                node.load_avg[0] if len(node.load_avg) > 0 else 0,
                node.load_avg[1] if len(node.load_avg) > 1 else 0,
                node.load_avg[2] if len(node.load_avg) > 2 else 0,
                node.uptime, node.system, node.processes,
                '是' if node.is_abnormal else '',
                ', '.join(node.abnormal_reasons) if node.abnormal_reasons else ''
            ]
            fill = (red_fill if node.is_abnormal
                    else yel_fill if node.status != 'online'
                    else None)
            for col, val in enumerate(vals, 1):
                cell = ws2.cell(row=row, column=col, value=val)
                if fill:
                    cell.fill = fill

        # 🔧 修复：使用 openpyxl 列字母工具，支持超过 26 列
        from openpyxl.utils import get_column_letter
        for col in range(1, 16):
            ws2.column_dimensions[get_column_letter(col)].width = 18

        wb.save(output_path)
        log.info(f'[Inspection] 报告已导出: {output_path}')


inspection_engine = InspectionEngine()